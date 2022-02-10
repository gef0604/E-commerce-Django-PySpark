import json

import yaml
import findspark
import pandas as pd
from pandas import ExcelWriter
findspark.init()
from pyspark.conf import SparkConf
from pyspark.sql import SparkSession
from pyspark import AccumulatorParam
from openpyxl import load_workbook
from collections import namedtuple
import helper
CONF_PATH = "conf.yml"
EXCEL_PATH = "test.xlsx"


class DataJobAbstract:
    """
    This method is to get conf as dict from conf file like yml
    """

    def build_conf(self):
        pass

    """
    This method is to get data as a list, like a list of dataframe
    """

    def get_datasets(self):
        pass

    """
    This method is to execute the core logic, like recon
    """

    def execute(self):
        pass

    """
    This is a post method to deal with the result of the execution, like send email, export as excel etc
    """

    def post(self, args):
        pass


class SparkDataFrameJob(DataJobAbstract):
    def __init__(self):
        self.spark = self.get_spark_session()

    def get_spark_session(self):
        sparkSession = (SparkSession
                        .builder
                        .appName('example-pyspark-read-and-write-from-hive')
                        .config("hive.metastore.uris", "thrift://localhost:9083", conf=SparkConf())
                        .enableHiveSupport()
                        .getOrCreate()
                        )

        return sparkSession


class SparkDataReconJob(SparkDataFrameJob):
    def test(self):
        df1 = self.spark.createDataFrame(data=[(1, 2), (3, 4)], schema=['a', 'b'])
        df2 = self.spark.createDataFrame(data=[(1, 2), (5, 6)], schema=['a', 'b'])
        df3 = df1.intersect(df2)
        df3.show()

    def test_execute(self):
        df1 = self.spark.createDataFrame(data=[(1, 2), (3, 4)], schema=['a', 'b'])
        df2 = self.spark.createDataFrame(data=[(1, 2), (3, 4)], schema=['a', 'b'])
        res = self.field_to_field_compare(df1,df2,'a')
        self.post(res)
    def execute(self):
        conf = YMLConfBuilder().get_conf(CONF_PATH)

        for dataset in conf['datasets'].keys():
            sqls = [conf['datasets'][dataset]['source1_sql'],conf['datasets'][dataset]['source2_sql']]
            l = SparkDataFrameFetcher().fetch(self.spark, sqls)
            # audit = self.audit(l[0],l[1])
            res = self.field_to_field_compare(l[0], l[1], 'user_id')
            self.post(res)
    # field to field compare
    def field_to_field_compare(self, df1, df2, key):
        """
        (key1,[(row,1) (row,2)])
        """
        def unite_same_field(x):
            # key = x[0]
            rows = x[1]
            count_cols = len(rows[0][0])
            res = tuple()
            for i in range(count_cols):
                # if only 1 row
                if len(rows) == 1:
                    if rows[0][1] == 1:
                        res = res + (rows[0][0][i], None,)
                    else:
                        res = res + (None, rows[0][0][i],)
                else:
                    if rows[0][1] == 1:
                        res = res + (rows[0][0][i], rows[1][0][i],)
                    else:
                        res = res + (rows[1][0][i], rows[0][0][i],)
            return res
        def count_mismatch(x,acc):
            column_names = column_names_str.split(",")
            # if key missing, which means the record not exist
            # then skip the mismatch
            # we count it as
            if x[2 * key_index] is None or x[2 * key_index + 1] is None:
                return 1

            for idx in range(len(column_names)):
                # if either of them don have primary key, continue, that means it's missing
                # should we throw exception?
                #
                # if x[2*idx] is None or x[2*idx + 1] == None:
                #     if idx != key_index:
                #         acc.add(column_names[idx])
                # elif x[2*idx] != x[2*idx + 1]:
                #     acc.add(column_names[idx])
                if x[2 * idx] != x[2 * idx + 1]:
                    acc.add(column_names[idx])
            # acc.add("qwe")
            return 1
        # case sensitive or not, need to deal with this to avoid bug

        # count missing key
        df_missing_from_src1 = df2.select(key).subtract(df1.select(key))
        df_missing_from_src2 = df1.select(key).subtract(df2.select(key))
        count_missing_sr1 = df_missing_from_src1.count()
        count_missing_sr2 = df_missing_from_src2.count()
        print("missing from src1:")
        df_missing_from_src1.show()
        print("missing from src2:")
        df_missing_from_src2.show()

        # change (field1,...key,..fieldn) => (key, (field1,...)), so we can combine key
        # union to dataset together so we can group by key
        # after map: (key,iterable[RowWithSameKey])
        key_index = df1.columns.index(key)
        rdd1 = df1.rdd.map(lambda x : (x[key_index],(x, 1)))
        rdd2 = df2.rdd.map(lambda x : (x[key_index],(x, 2)))
        union_rdd = rdd1.union(rdd2)
        grouped_union_rdd = union_rdd.groupByKey().map(lambda x : (x[0], list(x[1])))
        # print("example after group by key:")
        # print(grouped_union_rdd.collect()[0])

        # this rdd is: field1_src1, field1_src2, ....fieldx_src1, fieldx_src2
        # this is what we need in field to field compare sheet
        aggr_rdd = grouped_union_rdd.map(lambda x : helper.unite_same_field(x)).cache()

        # print("example after we put the same fields together:")
        # print(aggr_rdd)
        # aggr_df = self.spark.createDataFrame(data=aggr_rdd.collect(), schema=["_{}".format(str(i+1)) for i in range(2*len(df1.columns))]).toPandas()
        # aggr_df = aggr_rdd.toDF(["_{}".format(str(i+1)) for i in range(2*len(df1.columns))]).toPandas()
        # print(aggr_df)

        # defined a accumulator to count mismatch
        acc = self.spark.sparkContext.accumulator('initial', DictAccumulatorParam())
        acc.value = json.dumps({}, sort_keys=True)

        column_names_str = ",".join(df1.columns)

        # count mismatch by the field to field rdd
        # filterRDD is less confusing, this is just an action to put data in accumulator
        # we won't use the count_mismatch_rdd.count()
        count_mismatch_rdd = aggr_rdd.map(lambda x : count_mismatch(x,acc))
        count_mismatch_rdd.count()

        res = json.loads(acc.value)
        # this map is divided by partition, need to merge
        # example: {'{"b": 1}': 1, '{"a":1}': 5} => {'a':5, 'b':1}
        print(res)
        sum_up_count = {}
        for item in res.items():
            sub_res = json.loads(item[0])

            for sub_item in sub_res.items():
                if sub_item[0] not in sum_up_count.keys():
                    sum_up_count[sub_item[0]] = 0
                sum_up_count[sub_item[0]] = sum_up_count[sub_item[0]] + sub_item[1] * item[1]
        print(sum_up_count)

        audit = []

        for field in df1.columns:
            if field in sum_up_count.keys():
                audit.append({"Column": field,
                              "Mismatch": sum_up_count[field],
                              "Missing from src1": count_missing_sr1,
                              "Missing from src2": count_missing_sr2})
            else:
                audit.append({"Column": field,
                              "Mismatch": 0,
                              "Missing from src1": count_missing_sr1,
                              "Missing from src2": count_missing_sr2})

        df_audit = pd.DataFrame(data=audit)
        # aggr_df = self.spark.createDataFrame(data=aggr_rdd.collect(),
        #                                      schema=["_{}".format(str(i + 1)) for i in
        #                                              range(2 * len(df1.columns))]).toPandas()
        # df_audit.to_excel(EXCEL_PATH, sheet_name='audit')
        try:
            aggr_df = aggr_rdd.toDF().toPandas()
        except:
            aggr_df = self.spark.createDataFrame(data=aggr_rdd.collect(),
                                                 schema=["_{}".format(str(i + 1)) for i in range(2 * len(df1.columns))]).toPandas()
            print(aggr_df)
        return (aggr_df,df1.columns, df_audit)




        # print(rdd3.count())

    # audit, count mismatch
    def audit(self,df1, df2):
        acc = self.spark.sparkContext.accumulator('initial', DictAccumulatorParam())
        acc.value = json.dumps({}, sort_keys=True)
        acc.add("aer")
        acc.add("aer")
        print(acc)

        df1 = self.spark.createDataFrame(schema=['a','b'],data=[(1,2),(2,2)])
        df2 = self.spark.createDataFrame(schema=['a', 'b'], data=[(1, 2), (2, 3)])

        combined_rdd = df1.rdd.union(df2.rdd)
        key_index = 0





    def post(self, args):
        df_field_to_field = args[0]
        columns = args[1]
        df_audit = args[2]
        spanned_columns = []
        writer = ExcelWriter(EXCEL_PATH)
        for col in columns:
            spanned_columns.append(col)
            spanned_columns.append("")
        df_field_to_field.to_excel(writer,index=False, header=spanned_columns,sheet_name="field_to_field")
        df_audit.to_excel(writer,sheet_name="audit",index=False)
        writer.save()


class ConfBuilder:
    def get_conf(self, conf_path):
        pass


class YMLConfBuilder(ConfBuilder):
    def get_conf(self, conf_path):
        f = open(CONF_PATH)
        return yaml.load(f.read(), Loader=yaml.FullLoader)


class Fetcher:
    def fetch(self, *src):
        pass


class SparkDataFrameFetcher(Fetcher):
    def fetch(self, spark_session, src):
        print(src)
        res = []
        for sql in src:
            print(sql)
            res.append(spark_session.sql(sql))
        return res

class ExcelFormatter:
    def format_field_to_field_excel(self, path):
        pass

class SparkDataReconExcelFormatter(ExcelFormatter):
    def format_field_to_field_excel(self, path):
        wb = load_workbook(path)
        field_to_field = wb['Sheet1']
        # insert a row at before line 2
        """
        expected:
        fieldName1
        src1, src2
        val1, val2
        """
        field_to_field.insert_rows(2)

        # at second row, insert source1, source2
        for i in range(field_to_field.max_column):
            field_to_field.cell(row=2, column=i+1, value="source1" if i % 2 == 0 else "source2")

        # merge every two cells at the top
        for i in range(int(field_to_field.max_column/2)):
            field_to_field.merge_cells(start_row=1, start_column=2*i+1, end_row=1, end_column=2*i+2)
        wb.save(EXCEL_PATH)

class DictAccumulatorParam(AccumulatorParam):
    """
    maintain a countmap, to count the session steps
    """
    def zero(self, value):

        return '{}'

    def addInPlace(self, value1, value2):
        value1 = json.loads(value1)

        if value2 in value1.keys():
            value1[value2] = value1[value2] + 1

        else:
            value1[value2] = 1

        return json.dumps(value1, sort_keys=True)



class Mock:

    def generate(self):
        mocked_model = namedtuple("mocked", ['a', 'b'])
        src1 = [mocked_model(1,2), mocked_model(3,4)]
        src2 = [mocked_model(1,3), mocked_model(5,6)]

        # insert hive, two different table
        sparkSession = (SparkSession
                        .builder
                        .appName('example-pyspark-read-and-write-from-hive')
                        .config("hive.metastore.uris", "thrift://localhost:9083", conf=SparkConf())
                        .enableHiveSupport()
                        .getOrCreate()
                        )

        df1 = sparkSession.createDataFrame(src1)
        df2 = sparkSession.createDataFrame(src2)

        sparkSession.sql("DROP TABLE IF EXISTS test1")
        df1.write.saveAsTable("test1")

        sparkSession.sql("DROP TABLE IF EXISTS test2")
        df2.write.saveAsTable("test2")
    def verify(self):
        sparkSession = (SparkSession
                        .builder
                        .appName('example-pyspark-read-and-write-from-hive')
                        .config("hive.metastore.uris", "thrift://localhost:9083", conf=SparkConf())
                        .enableHiveSupport()
                        .getOrCreate()
                        )

        sparkSession.sql("select * from test1").show()
        sparkSession.sql("select * from test2").show()
class UnitTest:
    def test_spark_recon(self):
        sparkSession = (SparkSession
                        .builder
                        .appName('example-pyspark-read-and-write-from-hive')
                        .config("hive.metastore.uris", "thrift://localhost:9083", conf=SparkConf())
                        .enableHiveSupport()
                        .getOrCreate()
                        )

        df1 = sparkSession.sql("select * from test1")
        df2 = sparkSession.sql("select * from test2")

        recon = SparkDataReconJob()
        res = recon.field_to_field_compare(df1,df2,'a')
        recon.post(res)
# SparkDataReconJob().execute()
# SparkDataReconExcelFormatter().format_field_to_field_excel(EXCEL_PATH)
# SparkDataReconJob().audit(1,2)
# SparkDataReconJob().test()
# SparkDataReconJob().test_execute()
# mock = Mock()
# mock.generate()
# mock.verify()
UnitTest().test_spark_recon()